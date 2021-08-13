from NaverNewsCrawler import NaverNewsCrawler

# 저장하는 파일명에 대한 예외처리
import os

# json 처리 모듈
import json

####사용자로 부터 기사 수집을 원하는 키워드를 input을 이용해 입력받아 ? 부분에 넣으세요
keyword = input('수집할 기사 내용 검색: ')
crawler = NaverNewsCrawler(keyword)

#### 수집한 데이터를 저장할 엑셀 파일명을 input을 이용해 입력받아 ? 부분에 넣으세요
file_name = input('저장할 액셀 파일명: ')
file_name_extention = os.path.splitext(file_name)[-1] # 확장자 축출
file_name_base = os.path.basename(file_name).split('.') # 이름 축출
if file_name_extention != '.xlsx':     
    file_name = file_name_base[0] + '.xlsx'

crawler.get_news(file_name)

#### 아래코드를 실행해 이메일 발송 기능에 필요한 모듈을 임포트하세요.
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
import re

#### gmail 발송 기능에 필요한 계정 정보를 아래 코드에 입력하세요.
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
with open('conf.json', encoding='utf-8') as myUserInfo:
    config = json.load(myUserInfo)
SMTP_USER = config['email']
SMTP_PASSWORD = config['password']

#### 아래 코드를 실행해 메일 발송에 필요한 send_mail 함수를 만드세요.
def send_mail(name, addr, subject, contents, attachment=None):
    patt = '(^[a-zA-Z0-9_.-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)'
    #patt = '([a-zA-Z0-9]+)([\_\.\-{1}])?([a-zA-Z0-9]+)\@([a-zA-Z0-9]+)([\.])([a-zA-Z\.]+)'
    if not re.match(patt, addr):
        print('Wrong email')
        return

    msg = MIMEMultipart('alternative')
    if attachment:
        msg = MIMEMultipart('mixed')

    msg['From'] = SMTP_USER
    msg['To'] = addr
    msg['Subject'] = name + '님, ' + subject

    text = MIMEText(contents, _charset='utf-8')
    msg.attach(text)

    if attachment:
        from email.mime.base import MIMEBase
        from email import encoders

        file_data = MIMEBase('application', 'octect-stream')
        file_data.set_payload(open(attachment, 'rb').read())
        encoders.encode_base64(file_data)

        import os
        filename = os.path.basename(attachment)
        file_data.add_header('Content-Disposition', 'attachment; filename="' + filename + '"')
        msg.attach(file_data)

    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    smtp.close()

#### 프로젝트 폴더에 있는 email_list.xlsx 파일에 이메일 받을 사람들의 정보를 입력하세요.
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
with open('sendemail.json', encoding='utf-8') as senduser:
    send_user = json.load(senduser)
ws['A1'] = send_user['emailusername1'] # 보낼 이름 작성
ws['B1'] = send_user['email1'] # 보낼 메일 작성
ws['A2'] = send_user['emailusername2'] # 보낼 이름 작성
ws['B2'] = send_user['email2'] # 보낼 메일 작성

wb.save('email_list.xlsx')

#### 엑셀 파일의 정보를 읽어올 수 있는 모듈을 import하세요.
from openpyxl import load_workbook

#### email_list.xlsx 파일을 읽어와 해당 사람들에게 수집한 뉴스 정보 엑셀 파일을 send_mail 함수를 이용해 전송하세요.
wb = load_workbook('email_list.xlsx', read_only=True)
data = wb.active

subjects = keyword + '의 관련 뉴스 메일 입니다.'
contents = keyword + '''의 관련 뉴스 입니다.'''

for row in data.iter_rows():
    send_mail(row[0].value, row[1].value, subjects, contents, file_name)